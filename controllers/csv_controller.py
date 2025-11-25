from services.csv_service import CSVService, CSVServiceError
from models.csv_model import CSVData
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime, timedelta, time
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import io

class CSVContext:
    def __init__(self):
        self.data: CSVData | None = None
        self.device_columns: Dict = {}
        self.analysis_cache: Dict = {}
        self.device_configs: Dict[str, Dict[str, Any]] = {}

class CSVController:
    def __init__(self):
        print("--- CONTROLADOR V14: FIX EXPORT ---")
        self.contexts: Dict[str, CSVContext] = {
            'hora_exacta': CSVContext(),
            'general': CSVContext(),
            'ciclos': CSVContext(),
            'escalones': CSVContext()
        }
        self.last_warning: str | None = None
        self.VOLTAGE = 120.0 

    # =========================================================================
    #  MÉTODOS DE CÁLCULO
    # =========================================================================
    def get_monthly_projection(self) -> Tuple[List[Dict], float]:
        rows, totals = self.get_energy_summary()
        monthly_rows = []
        grand_total_month = 0.0
        
        # Agrupar y calcular
        temp_list = []
        luminarias_total = 0.0
        found_luminarias = False
        
        for r in rows:
            month_kwh = r['total_week'] * 4
            grand_total_month += month_kwh
            name = r['device']
            
            if "luminaria" in name.lower() or "iluminacion" in name.lower() or "iluminación" in name.lower():
                luminarias_total += month_kwh
                found_luminarias = True
            else:
                temp_list.append({'device': name, 'kwh_month': month_kwh})
        
        if found_luminarias:
            temp_list.append({'device': 'Iluminación (Agrupada)', 'kwh_month': luminarias_total})
            
        temp_list.sort(key=lambda x: x['kwh_month'], reverse=True)
        
        final_rows = []
        accumulated_kwh = 0.0
        
        for item in temp_list:
            kwh = item['kwh_month']
            accumulated_kwh += kwh
            rel_energy = (kwh / grand_total_month * 100) if grand_total_month > 0 else 0.0
            acc_rel = (accumulated_kwh / grand_total_month * 100) if grand_total_month > 0 else 0.0
            
            final_rows.append({
                'device': item['device'],
                'kwh_month': round(kwh, 4),
                'rel_energy': round(rel_energy, 2),
                'acc_kwh': round(accumulated_kwh, 4),
                'acc_rel': round(acc_rel, 2)
            })
        return final_rows, round(grand_total_month, 4)

    def get_energy_summary(self) -> Tuple[List[Dict], Dict]:
        summary_rows = []
        grand_totals = {'daily_wd': 0.0, 'daily_we': 0.0, 'total_5d': 0.0, 'total_2d': 0.0, 'total_week': 0.0}
        factor = 1.0 / 60000.0

        for ctx in ['hora_exacta', 'ciclos', 'escalones']:
            devices = self.get_devices(ctx)
            for dev in devices:
                _, p_wd = self.get_typical_day_profile(ctx, dev, 'weekday')
                _, p_we = self.get_typical_day_profile(ctx, dev, 'weekend')
                
                kwh_day_wd = sum(p_wd) * factor
                kwh_day_we = sum(p_we) * factor
                total_5d = kwh_day_wd * 5
                total_2d = kwh_day_we * 2
                total_week = total_5d + total_2d
                
                summary_rows.append({
                    'section': ctx.replace('_', ' ').title(),
                    'device': dev,
                    'daily_wd': round(kwh_day_wd, 4),
                    'daily_we': round(kwh_day_we, 4),
                    'total_5d': round(total_5d, 4),
                    'total_2d': round(total_2d, 4),
                    'total_week': round(total_week, 4)
                })
                grand_totals['daily_wd'] += kwh_day_wd
                grand_totals['daily_we'] += kwh_day_we
                grand_totals['total_5d'] += total_5d
                grand_totals['total_2d'] += total_2d
                grand_totals['total_week'] += total_week
        
        for k in grand_totals: grand_totals[k] = round(grand_totals[k], 4)
        return summary_rows, grand_totals

    # --- GESTIÓN DE MEMORIA ---
    def set_device_config_simple(self, context_key, device_name, count, starts, ends=None):
        if context_key in self.contexts:
            self.contexts[context_key].device_configs[device_name] = {
                'type': 'simple', 'count': count, 'starts': starts, 'ends': ends or []
            }

    def set_device_config_weekly(self, context_key, device_name, wd_count, wd_starts, wd_ends, we_count, we_starts, we_ends):
        if context_key in self.contexts:
            self.contexts[context_key].device_configs[device_name] = {
                'type': 'weekly',
                'weekday': {'count': wd_count, 'starts': wd_starts, 'ends': wd_ends},
                'weekend': {'count': we_count, 'starts': we_starts, 'ends': we_ends}
            }

    def get_device_config(self, context_key, device_name):
        if context_key in self.contexts:
            return self.contexts[context_key].device_configs.get(device_name, {})
        return {}

    # --- LECTURA ---
    def load_csv(self, path: str, context_key: str):
        if context_key not in self.contexts: self.contexts[context_key] = CSVContext()
        ctx = self.contexts[context_key]
        try:
            ctx.data = CSVService.read_csv(path)
            ctx.analysis_cache.clear()
            ctx.device_configs.clear()
        except CSVServiceError: raise
        except Exception as e: raise CSVServiceError(f"Error inesperado al leer CSV: {e}")

        if not ctx.data.columns: raise CSVServiceError("CSV sin encabezados.")
        if len(ctx.data.columns) < 1: raise CSVServiceError("El CSV está vacío.")

        self._parse_device_pairs(ctx, context_key)
        if not ctx.device_columns: raise CSVServiceError("No se encontraron dispositivos válidos.")
        return ctx.data

    def _parse_device_pairs(self, ctx: CSVContext, context_key: str):
        ctx.device_columns = {}
        cols = [col.strip() for col in ctx.data.columns]
        i = 0
        pairs_found = False
        while i + 1 < len(cols):
            fecha_col = cols[i]
            value_col = cols[i + 1]
            if "fecha" in fecha_col.lower() or "hora" in fecha_col.lower() or "time" in fecha_col.lower():
                device_name = self._extract_device_name(fecha_col)
                if not device_name: device_name = value_col
                device_name = device_name.strip()
                if device_name:
                    original = device_name
                    suffix = 1
                    while device_name in ctx.device_columns:
                        suffix += 1
                        device_name = f"{original}_{suffix}"
                    ctx.device_columns[device_name] = (fecha_col, value_col)
                    pairs_found = True
            i += 2

        if context_key == 'escalones' and not pairs_found:
            for col in cols:
                if not col: continue
                device_name = col.strip()
                original = device_name
                suffix = 1
                while device_name in ctx.device_columns:
                    suffix += 1
                    device_name = f"{original}_{suffix}"
                ctx.device_columns[device_name] = (None, col)

    def _extract_device_name(self, date_column: str) -> str:
        clean_name = date_column
        patterns = ['fecha hora', 'fecha/hora', 'fechahora', 'fecha', 'hora', 'timestamp']
        for p in patterns: clean_name = clean_name.lower().replace(p, '')
        clean_name = ' '.join(clean_name.split()).strip()
        if clean_name and clean_name.islower(): clean_name = clean_name.title()
        return clean_name if clean_name else ""

    def get_devices(self, context_key: str):
        if context_key in self.contexts: return list(self.contexts[context_key].device_columns.keys())
        return []

    def _parse_date(self, date_str: str) -> Optional[datetime]:
        if not date_str: return None
        formats = ["%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]
        for fmt in formats:
            try: return datetime.strptime(date_str.strip(), fmt)
            except ValueError: continue
        return None

    # --- TABLA DUAL ---
    def get_dual_table_data(self, context_key: str, device_name: str) -> List[Tuple[str, str, str]]:
        if context_key == 'hora_exacta':
            rows = self.get_values_for_device(context_key, device_name)
            return [(r[0], r[1], r[1]) for r in rows]

        config = self.get_device_config(context_key, device_name)
        wd_starts, wd_ends = [], []
        we_starts, we_ends = [], []

        if config.get('type') == 'weekly':
            wd = config.get('weekday', {})
            we = config.get('weekend', {})
            wd_starts, wd_ends = wd.get('starts', []), wd.get('ends', [])
            we_starts, we_ends = we.get('starts', []), we.get('ends', [])
        else:
            start = config.get('starts', ["00:00"])
            end = config.get('ends', ["01:00"])
            wd_starts, wd_ends = start, end
            we_starts, we_ends = start, end

        rows_wd = self.get_values_for_device(context_key, device_name, wd_starts, wd_ends)
        rows_we = self.get_values_for_device(context_key, device_name, we_starts, we_ends)

        combined = []
        max_len = max(len(rows_wd), len(rows_we))
        for i in range(max_len):
            t_str = rows_wd[i][0] if i < len(rows_wd) else (rows_we[i][0] if i < len(rows_we) else "")
            val_wd = rows_wd[i][1] if i < len(rows_wd) else "0"
            val_we = rows_we[i][1] if i < len(rows_we) else "0"
            combined.append((t_str, val_wd, val_we))
        return combined

    # --- OBTENCIÓN DE DATOS ---
    def get_values_for_device(self, context_key: str, device_name: str, start_times: List[str] = None, end_times: List[str] = None) -> List[Tuple[str, str]]:
        self.last_warning = None
        if context_key not in self.contexts or not self.contexts[context_key].data:
            raise CSVServiceError(f"No hay datos cargados en {context_key}.")
        ctx = self.contexts[context_key]
        if device_name not in ctx.device_columns: raise CSVServiceError(f"Dispositivo '{device_name}' no encontrado.")
        fecha_col, val_col = ctx.device_columns[device_name]
        raw_data = []
        nominal_power_str = "0"
        max_val_found = 0.0

        if fecha_col:
            try:
                fecha_idx = ctx.data.columns.index(fecha_col)
                val_idx = ctx.data.columns.index(val_col)
            except ValueError: raise CSVServiceError("Error de índices.")
            for row in ctx.data.rows:
                f_str = row[fecha_idx] if fecha_idx < len(row) else ""
                v_str = row[val_idx] if val_idx < len(row) else ""
                dt = self._parse_date(f_str)
                if dt: raw_data.append((dt, f_str, v_str))
                try:
                    val = float(v_str.replace(',', '.'))
                    if val > max_val_found:
                        max_val_found = val
                        nominal_power_str = v_str
                except: continue
            raw_data.sort(key=lambda x: x[0])
        else:
            try: val_idx = ctx.data.columns.index(val_col)
            except ValueError: raise CSVServiceError("Error de índices.")
            for row in ctx.data.rows:
                if val_idx < len(row):
                    v_str = row[val_idx]
                    try:
                        val = float(v_str.replace(',', '.'))
                        if val > max_val_found:
                            max_val_found = val
                            nominal_power_str = v_str
                    except: continue

        if context_key == 'hora_exacta' and "nevera" in device_name.lower():
            final_data = self._process_nevera_logic(raw_data)
            return [(item[1], item[2]) for item in final_data]
        elif context_key == 'ciclos' and start_times is not None:
            multi_cycle = self._apply_multi_cycle_day(raw_data, start_times)
            return [(item[1], item[2]) for item in multi_cycle]
        elif context_key == 'escalones' and start_times is not None and end_times is not None:
            if raw_data: base_date = raw_data[0][0].date()
            else: base_date = datetime.now().date()
            step_data = self._generate_step_profile(nominal_power_str, base_date, start_times, end_times)
            return [(item[1], item[2]) for item in step_data]
        else:
            return [(item[1], item[2]) for item in raw_data]

    # --- VECTORES ---
    def get_daily_power_vector(self, context_key: str, device_name: str, starts=None, ends=None) -> List[float]:
        data_rows = self.get_values_for_device(context_key, device_name, starts, ends)
        if not data_rows: return [0.0] * 1440
        power_axis = [0.0] * 1440
        try:
            first_dt = datetime.strptime(data_rows[0][0], "%d/%m/%Y %H:%M:%S")
            start_of_day = datetime.combine(first_dt.date(), time(0,0))
        except: start_of_day = datetime.combine(datetime.now().date(), time(0,0))
        minute_buckets = {i: [] for i in range(1440)}
        for t_str, v_str in data_rows:
            try:
                dt = datetime.strptime(t_str, "%d/%m/%Y %H:%M:%S")
                minute_idx = int((dt - start_of_day).total_seconds() // 60)
                minute_idx = minute_idx % 1440
                val = float(v_str.replace(',', '.'))
                minute_buckets[minute_idx].append(val)
            except: continue
        for i in range(1440):
            values = minute_buckets[i]
            if values:
                avg = sum(values) / len(values)
                if context_key == 'escalones': power_axis[i] = avg
                else: power_axis[i] = avg * self.VOLTAGE
        return power_axis

    def get_typical_day_profile(self, context_key: str, device_name: str, day_type: str) -> Tuple[List[datetime], List[float]]:
        config = self.get_device_config(context_key, device_name)
        starts, ends = [], []
        if config.get('type') == 'weekly':
            sub = config.get(day_type, {})
            starts, ends = sub.get('starts', []), sub.get('ends', [])
        else:
            starts, ends = config.get('starts'), config.get('ends')
        p_vec = self.get_daily_power_vector(context_key, device_name, starts, ends)
        base = datetime.now().date()
        t_axis = [datetime.combine(base, time(0,0)) + timedelta(minutes=i) for i in range(1440)]
        return t_axis, p_vec

    def get_total_typical_profile(self, day_type: str, is_energy=False) -> Tuple[List[datetime], List[float]]:
        total = [0.0] * 1440
        time_axis = []
        for ctx in ['hora_exacta', 'ciclos', 'escalones']:
            for dev in self.get_devices(ctx):
                t, p = self.get_typical_day_profile(ctx, dev, day_type)
                if not time_axis: time_axis = t
                for i in range(1440):
                    if i < len(p): total[i] += p[i]
        if not time_axis:
            base = datetime.now().date()
            time_axis = [datetime.combine(base, time(0,0)) + timedelta(minutes=i) for i in range(1440)]
        if is_energy:
            eng = [0.0] * 1440
            acc = 0.0
            for i in range(1440):
                acc += total[i] * (1.0/60000.0)
                eng[i] = acc
            return time_axis, eng
        return time_axis, total

    def get_weekly_power_vector(self, context_key: str, device_name: str) -> Tuple[List[str], List[float]]:
        t_wd, p_wd = self.get_typical_day_profile(context_key, device_name, 'weekday')
        t_we, p_we = self.get_typical_day_profile(context_key, device_name, 'weekend')
        w_p, w_t = [], []
        base = datetime.now().date()
        base = base - timedelta(days=base.weekday())
        curr = datetime.combine(base, time(0,0))
        for d in range(7):
            prof = p_wd if d < 5 else p_we
            w_p.extend(prof)
            for _ in range(1440):
                w_t.append(curr)
                curr += timedelta(minutes=1)
        return w_t, w_p

    def get_total_weekly_vector(self, is_energy=False) -> Tuple[List[str], List[float]]:
        tot = [0.0] * 10080
        tm = []
        for ctx in ['hora_exacta', 'ciclos', 'escalones']:
            for dev in self.get_devices(ctx):
                t, p = self.get_weekly_power_vector(ctx, dev)
                if not tm: tm = t
                for i in range(min(len(tot), len(p))): tot[i] += p[i]
        if not tm:
            base = datetime.now().date()
            base = base - timedelta(days=base.weekday())
            curr = datetime.combine(base, time(0,0))
            for _ in range(10080):
                tm.append(curr)
                curr += timedelta(minutes=1)
        if is_energy:
            eng = [0.0]*10080
            acc = 0.0
            for i in range(10080):
                acc += tot[i] * (1.0/60000.0)
                eng[i] = acc
            return tm, eng
        return tm, tot

    # --- HELPERS INTERNOS ---
    def get_power_profile_1min(self, k, d): return self.get_typical_day_profile(k, d, 'weekday')
    def get_total_power_vector(self): return self.get_total_typical_profile('weekday')
    def get_energy_profile(self, k, d):
        t, p = self.get_typical_day_profile(k, d, 'weekday')
        e, acc = [], 0
        for v in p:
            acc += v*(1.0/60000.0)
            e.append(acc)
        return t, e
    def get_total_energy_vector(self): return self.get_total_typical_profile('weekday', True)

    def _generate_step_profile(self, nominal_val_str, base_date, start_times, end_times):
        timeline = []
        current = datetime.combine(base_date, time(0,0))
        end_of_day = current + timedelta(hours=24)
        while current < end_of_day:
            timeline.append({'dt': current, 'str': current.strftime("%d/%m/%Y %H:%M:%S"), 'val': "0"})
            current += timedelta(minutes=1)
        if not start_times: return [(t['dt'], t['str'], t['val']) for t in timeline]
        for i in range(len(start_times)):
            if i >= len(end_times): break
            try:
                t_s = datetime.strptime(start_times[i], "%H:%M").time()
                t_e = datetime.strptime(end_times[i], "%H:%M").time()
                dt_s = datetime.combine(base_date, t_s)
                dt_e = datetime.combine(base_date, t_e)
                if dt_e < dt_s:
                    dt_end_day = datetime.combine(base_date, time(23,59,59))
                    dt_start_day = datetime.combine(base_date, time(0,0))
                    for p in timeline:
                        if dt_s <= p['dt'] <= dt_end_day: p['val'] = nominal_val_str
                        if dt_start_day <= p['dt'] < dt_e: p['val'] = nominal_val_str
                else:
                    for p in timeline:
                        if dt_s <= p['dt'] < dt_e: p['val'] = nominal_val_str
            except: continue
        return [(t['dt'], t['str'], t['val']) for t in timeline]

    def _apply_multi_cycle_day(self, raw_data, start_times_str):
        if not raw_data: return []
        target_times = []
        for t in start_times_str:
            try:
                try: tt = datetime.strptime(t, "%H:%M").time()
                except: tt = datetime.strptime(t, "%H:%M:%S").time()
                target_times.append(tt)
            except: continue
        base = raw_data[0][0].date()
        day_s = datetime.combine(base, time(0,0))
        day_e = day_s + timedelta(hours=24)
        if not target_times:
            zeros = []
            curr = day_s
            while curr < day_e:
                zeros.append((curr, curr.strftime("%d/%m/%Y %H:%M:%S"), "0"))
                curr += timedelta(minutes=1)
            return zeros
        target_times.sort()
        cycle_dur = raw_data[-1][0] - raw_data[0][0]
        orig_first = raw_data[0][0]
        active_ranges = []
        for t in target_times:
            start = datetime.combine(base, t)
            end = start + cycle_dur
            if end > day_e:
                active_ranges.append((start, day_e))
                active_ranges.append((day_s, day_s + (end - day_e)))
            else:
                active_ranges.append((start, end))
        final_rows = []
        curr = day_s
        while curr < day_e:
            is_active = False
            for s, e in active_ranges:
                if s <= curr <= e: is_active = True; break
            if not is_active: final_rows.append((curr, curr.strftime("%d/%m/%Y %H:%M:%S"), "0"))
            curr += timedelta(minutes=1)
        for t in target_times:
            cycle_start = datetime.combine(base, t)
            offset = cycle_start - orig_first
            for dt, _, val in raw_data:
                new_dt = dt + offset
                while new_dt >= day_e: new_dt -= timedelta(hours=24)
                while new_dt < day_s: new_dt += timedelta(hours=24)
                final_rows.append((new_dt, new_dt.strftime("%d/%m/%Y %H:%M:%S"), val))
        final_rows.sort(key=lambda x: x[0])
        return final_rows

    def _process_nevera_logic(self, sorted_data):
        if not sorted_data: return []
        s = sorted_data[0][0]
        e = sorted_data[-1][0]
        if (e-s) < timedelta(hours=24): pass
        target = (s + timedelta(days=1)).date()
        syn = []
        for dt, _, v in sorted_data:
            if dt.date() == target: syn.append((dt, dt.strftime("%d/%m/%Y %H:%M:%S"), v))
            elif dt.date() == s.date():
                s_dt = dt + timedelta(days=1)
                if s_dt.date() == target: syn.append((s_dt, s_dt.strftime("%d/%m/%Y %H:%M:%S"), v))
        syn.sort(key=lambda x: x[0])
        return [x for x in syn if x[0].date() == target]

    def get_device_statistics(self, context_key: str, device_name: str) -> Dict:
        return {}
    def get_all_statistics(self, context_key: str) -> Dict:
        return {}
    
    # ========================================================
    #  EXPORTACIÓN A EXCEL
    # ========================================================
    def export_report(self, filename: str, figures: Dict[str, Any] = None, bill_real: float = 0.0):
        import pandas as pd
        import openpyxl
        from openpyxl.drawing.image import Image as ExcelImage
        import io
        
        # 1. DATOS MINUTO A MINUTO (L-V y S-D)
        base_date = datetime.now().date()
        time_axis = [datetime.combine(base_date, time(0,0)) + timedelta(minutes=i) for i in range(1440)]
        str_time = [t.strftime("%H:%M") for t in time_axis]
        
        data_lv = {"Hora": str_time}
        data_sd = {"Hora": str_time}
        total_lv = [0.0] * 1440
        total_sd = [0.0] * 1440
        
        for ctx in ['hora_exacta', 'ciclos', 'escalones']:
            for dev in self.get_devices(ctx):
                _, p_wd = self.get_typical_day_profile(ctx, dev, 'weekday')
                _, p_we = self.get_typical_day_profile(ctx, dev, 'weekend')
                col = f"{dev} ({ctx}) [W]"
                data_lv[col] = p_wd
                data_sd[col] = p_we
                for i in range(1440):
                    total_lv[i] += p_wd[i]
                    total_sd[i] += p_we[i]
        
        data_lv["TOTAL [W]"] = total_lv
        data_sd["TOTAL [W]"] = total_sd
        df_lv = pd.DataFrame(data_lv)
        df_sd = pd.DataFrame(data_sd)

        # 2. RESUMEN SEMANAL
        rows_data, totals = self.get_energy_summary()
        df_weekly = pd.DataFrame(rows_data)
        df_weekly = df_weekly.rename(columns={
            'section': 'Sección', 'device': 'Dispositivo',
            'daily_wd': 'Día Laboral (kWh)', 'daily_we': 'Fin de Semana (kWh)',
            'total_5d': 'Total L-V (kWh)', 'total_2d': 'Total S-D (kWh)',
            'total_week': 'Total Semanal (kWh)'
        })
        total_row = {
            'Sección': '', 'Dispositivo': 'TOTAL GENERAL',
            'Día Laboral (kWh)': totals['daily_wd'], 'Fin de Semana (kWh)': totals['daily_we'],
            'Total L-V (kWh)': totals['total_5d'], 'Total S-D (kWh)': totals['total_2d'],
            'Total Semanal (kWh)': totals['total_week']
        }
        # Mapeo manual para evitar errores
        map_row = {
            'Sección': total_row['Sección'], 'Dispositivo': total_row['Dispositivo'],
            'Día Laboral (kWh)': total_row['Día Laboral (kWh)'], 
            'Fin de Semana (kWh)': total_row['Fin de Semana (kWh)'],
            'Total L-V (kWh)': total_row['Total L-V (kWh)'], 
            'Total S-D (kWh)': total_row['Total S-D (kWh)'],
            'Total Semanal (kWh)': total_row['Total Semanal (kWh)']
        }
        df_weekly = pd.concat([df_weekly, pd.DataFrame([map_row])], ignore_index=True)

        # 3. PROYECCIÓN MENSUAL
        rows_monthly, total_monthly = self.get_monthly_projection()
        df_monthly = pd.DataFrame(rows_monthly)
        df_monthly = df_monthly.rename(columns={
            'device': 'Dispositivo', 'kwh_month': 'Energía (kWh/mes)',
            'rel_energy': '% Relativo', 'acc_kwh': 'Acumulado (kWh)',
            'acc_rel': '% Acumulado'
        })
        cols_keep = ['Dispositivo', 'Energía (kWh/mes)', '% Relativo', 'Acumulado (kWh)', '% Acumulado']
        if all(c in df_monthly.columns for c in cols_keep): df_monthly = df_monthly[cols_keep]
        
        row_tot_month = {
            'Dispositivo': 'TOTAL GENERAL',
            'Energía (kWh/mes)': total_monthly,
            '% Relativo': '100%', 'Acumulado (kWh)': total_monthly, '% Acumulado': '100%'
        }
        df_monthly = pd.concat([df_monthly, pd.DataFrame([row_tot_month])], ignore_index=True)

        # 4. FACTURA
        diff = abs(total_monthly - bill_real)
        perc = (diff / bill_real * 100) if bill_real > 0 else 0.0
        status = "✅ Preciso" if perc <= 10 else ("⚠️ Aceptable" if perc <= 20 else "❌ Revisar")
        df_bill = pd.DataFrame({
            "Concepto": ["Energía Calculada (Mes)", "Energía Factura (Real)", "Diferencia (Absoluta)", "Desviación (%)", "Estado"],
            "Valor": [total_monthly, bill_real, diff, f"{perc:.2f}", status],
            "Unidad": ["kWh", "kWh", "kWh", "%", "-"]
        })

        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Escribir en el orden pedido
                df_lv.to_excel(writer, sheet_name='L-V Potencia', index=False)
                df_sd.to_excel(writer, sheet_name='S-D Potencia', index=False)
                df_weekly.to_excel(writer, sheet_name='Resumen Semanal', index=False)
                df_monthly.to_excel(writer, sheet_name='Proyección Mensual', index=False)
                df_bill.to_excel(writer, sheet_name='Comparativa de factura', index=False)
                
                for sheet_name in writer.sheets:
                    sheet = writer.sheets[sheet_name]
                    for column in sheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                            except: pass
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            if figures:
                wb = openpyxl.load_workbook(filename)
                ws = wb['Proyección Mensual']
                row_idx = len(df_monthly) + 4 
                for name, fig in figures.items():
                    if fig:
                        buf = io.BytesIO()
                        fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
                        buf.seek(0)
                        img = ExcelImage(buf)
                        img.anchor = f'B{row_idx}'
                        ws.add_image(img)
                        ws[f'B{row_idx-1}'] = name
                        ws[f'B{row_idx-1}'].font = openpyxl.styles.Font(bold=True)
                        row_idx += 25
                wb.save(filename)
        except Exception as e: raise CSVServiceError(f"Error escribiendo Excel: {e}")